import csv
import openpyxl
from io import BytesIO, StringIO
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.http import HttpResponse
from django.db.models import Q
import logging

from .models import (
    User, FeedbackSession, StudentSemester,
    Branch, Semester, SessionOffering
)
from .serializers import (
    StudentSemesterSerializer, FeedbackSessionSerializer
)

logger = logging.getLogger(__name__)

class StudentImportService:
    """
    Service layer for intelligently parsing and importing student data from Excel/CSV.
    Uses openpyxl and csv instead of pandas to save memory.
    """
    
    COLUMN_MAPPINGS = {
        'name': ['name', 'studentname', 'fullname', 'student_name', 'full_name'],
        'enroll_number': ['enrollnumber', 'enrollmentno', 'enrollno', 'rollno', 'enroll_number', 'enrollment_no', 'enrollmentnumber'],
        'department': ['department', 'dept', 'branch', 'dept_name', 'dept_code'],
        'semester': ['semester', 'sem', 'sem_no', 'term'],
        'session': ['session', 'academic_year', 'year', 'session_name', 'acad_year']
    }

    @staticmethod
    def normalize_header(header):
        """Normalize header string for matching."""
        if header is None:
            return ""
        return str(header).lower().strip().replace(' ', '').replace('_', '').replace('.', '')

    @classmethod
    def map_columns(cls, headers):
        """Maps data frame columns to required fields based on aliases."""
        mapped = {}
        normalized_cols = {cls.normalize_header(col): col for col in headers}
        
        for field, variants in cls.COLUMN_MAPPINGS.items():
            for variant in variants:
                norm_variant = cls.normalize_header(variant)
                if norm_variant in normalized_cols:
                    mapped[field] = normalized_cols[norm_variant]
                    break
        
        return mapped

    @classmethod
    def _detect_all_valid_sheets(cls, file_bytes, required_fields):
        """
        Scan every sheet in the workbook and return a list of
        (sheet_name, mapped_cols) for ALL sheets whose headers
        satisfy the required fields.
        """
        logs = []
        valid_sheets = []
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        all_sheets = wb.sheetnames
        logs.append(f"Workbook contains {len(all_sheets)} sheet(s): {', '.join(all_sheets)}")

        for name in all_sheets:
            try:
                sheet = wb[name]
                # Get the first row (headers)
                headers = []
                for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
                    headers.append(cell.value)
                
                mapped = cls.map_columns(headers)
                matched = [f for f in required_fields if f in mapped]
                missing = [f for f in required_fields if f not in mapped]

                if missing:
                    logs.append(
                        f"Sheet '{name}' -> skipped (matched {len(matched)}/{len(required_fields)}: "
                        f"{', '.join(matched)}; missing: {', '.join(missing)})"
                    )
                else:
                    logs.append(f"Sheet '{name}' -> VALID (all required columns found)")
                    valid_sheets.append((name, mapped))
            except Exception as exc:
                logs.append(f"Sheet '{name}' -> error reading headers: {exc}")

        return valid_sheets, logs

    @classmethod
    def _parse_rows(cls, row_iterator, mapped_cols, sheet_name, seen_enrollments, target_session_name=None):
        """
        Parse rows from an iterator of dictionaries.
        Returns (results_list, errors_list).
        """
        results = []
        errors = []

        for row_num, row in enumerate(row_iterator, start=2):
            try:
                name_val = row.get(mapped_cols.get('name'))
                enroll_val = row.get(mapped_cols.get('enroll_number'))
                dept_val = row.get(mapped_cols.get('department'))
                sem_val = row.get(mapped_cols.get('semester'))
                
                name = str(name_val).strip() if name_val is not None else None
                enroll_no = str(enroll_val).strip() if enroll_val is not None else None
                dept = str(dept_val).strip() if dept_val is not None else None
                sem = str(sem_val).strip() if sem_val is not None else None
                
                session = target_session_name
                session_col = mapped_cols.get('session')
                if session_col and row.get(session_col) is not None:
                    session = str(row[session_col]).strip()

                if not enroll_no:
                    errors.append(f"[{sheet_name}] Row {row_num} skipped: Missing Enrollment Number")
                    continue

                if enroll_no in seen_enrollments:
                    errors.append(f"[{sheet_name}] Row {row_num} skipped: Duplicate Enrollment Number '{enroll_no}'")
                    continue

                seen_enrollments.add(enroll_no)

                if not all([name, dept, sem, session]):
                    missing = [f for f, v in [('Name', name), ('Dept', dept), ('Sem', sem), ('Session', session)] if not v]
                    errors.append(f"[{sheet_name}] Row {row_num} (Enroll: {enroll_no}) skipped: Missing {', '.join(missing)}")
                    continue

                results.append({
                    'name': name,
                    'enrollment_no': enroll_no,
                    'department': dept,
                    'semester': sem,
                    'session': session,
                    'row_num': row_num,
                    'source_sheet': sheet_name
                })

            except Exception as e:
                errors.append(f"[{sheet_name}] Row {row_num} skipped: Unexpected error - {str(e)}")

        return results, errors

    @classmethod
    def process(cls, file, uploaded_by, preview=False, sheet_name=None, update_existing=True, session_id=None):
        required_fields = ['name', 'enroll_number', 'department', 'semester']
        target_session_name = None
        if session_id:
            try:
                target_session_name = FeedbackSession.objects.get(pk=session_id).name
            except FeedbackSession.DoesNotExist:
                return {
                    'success': False,
                    'error': f"Invalid session_id provided: {session_id}",
                    'sheet_logs': []
                }
        else:
            required_fields.append('session')
            
        sheet_logs = []

        try:
            if file.name.endswith('.csv'):
                content = file.read().decode('utf-8')
                csv_reader = csv.DictReader(StringIO(content))
                headers = csv_reader.fieldnames
                sheet_logs.append("CSV file detected -- reading single sheet.")
                
                mapped_cols = cls.map_columns(headers)
                missing_fields = [f for f in required_fields if f not in mapped_cols]
                if missing_fields:
                    return {
                        'success': False,
                        'error': f"Could not find required columns: {', '.join(missing_fields)}",
                        'mapped_columns': mapped_cols,
                        'available_columns': headers,
                        'sheet_logs': sheet_logs
                    }
                
                sheet_logs.append("Column mapping: " + ", ".join(f"{k} -> '{v}'" for k, v in mapped_cols.items()))
                all_results, all_errors = cls._parse_rows(csv_reader, mapped_cols, 'CSV', set(), target_session_name)
            
            else:
                file_bytes = file.read()
                wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
                
                if sheet_name:
                    if sheet_name not in wb.sheetnames:
                        return {
                            'success': False,
                            'error': f"Sheet '{sheet_name}' not found.",
                            'available_sheets': wb.sheetnames,
                            'sheet_logs': sheet_logs
                        }
                    valid_sheets = [(sheet_name, cls.map_columns([c.value for c in next(wb[sheet_name].iter_rows(min_row=1, max_row=1))]))]
                else:
                    valid_sheets, detect_logs = cls._detect_all_valid_sheets(file_bytes, required_fields)
                    sheet_logs.extend(detect_logs)

                if not valid_sheets:
                    return {
                        'success': False,
                        'error': "No valid student data found in any sheet.",
                        'available_sheets': wb.sheetnames,
                        'sheet_logs': sheet_logs
                    }

                all_results = []
                all_errors = []
                seen_enrollments = set()

                for sname, smapped in valid_sheets:
                    sheet = wb[sname]
                    # Convert sheet rows to list of dicts for _parse_rows
                    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
                    rows_iter = []
                    # Start from row 2
                    for row in sheet.iter_rows(min_row=2):
                        row_dict = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
                        rows_iter.append(row_dict)
                    
                    rows, errs = cls._parse_rows(rows_iter, smapped, sname, seen_enrollments, target_session_name)
                    all_results.extend(rows)
                    all_errors.extend(errs)
                    sheet_logs.append(f"Sheet '{sname}' -> {len(rows)} valid rows, {len(errs)} errors")

            if preview:
                return {
                    'success': True,
                    'preview': True,
                    'data': all_results,
                    'errors': all_errors,
                    'total_valid_rows': len(all_results),
                    'total_error_rows': len(all_errors),
                    'sheet_logs': sheet_logs
                }

            stats = cls.save_to_db(all_results, uploaded_by, update_existing)
            stats['errors'].extend(all_errors)
            return {
                'success': True,
                'preview': False,
                'stats': stats,
                'sheet_logs': sheet_logs
            }

        except Exception as e:
            logger.exception("Error processing student upload")
            return {
                'success': False,
                'error': f"Processing failed: {str(e)}",
                'sheet_logs': sheet_logs
            }

    @classmethod
    def save_to_db(cls, student_list, uploaded_by, update_existing=True):
        created_count = 0
        updated_count = 0
        skipped_existing = 0
        error_count = 0
        db_errors = []
        warnings = []
        processed_configs = set()
        
        with transaction.atomic():
            for data in student_list:
                enroll_no = data.get('enrollment_no', 'Unknown')
                source = data.get('source_sheet', '?')
                try:
                    name = data['name']
                    dept_code = data['department']
                    sem_val = data['semester']
                    session_name = data['session']
                    
                    session_obj = FeedbackSession.objects.filter(name__iexact=session_name).first()
                    if not session_obj:
                        db_errors.append(f"[{source}] Row {data['row_num']} (Enroll: {enroll_no}): Session '{session_name}' not found")
                        error_count += 1
                        continue
                    
                    branch_obj = Branch.objects.filter(Q(code__iexact=dept_code) | Q(name__iexact=dept_code)).first()
                    if not branch_obj:
                        branch_obj = Branch.objects.create(code=dept_code.upper()[:10], name=dept_code)
                    
                    sem_num = ''.join(filter(str.isdigit, str(sem_val)))
                    if not sem_num:
                        db_errors.append(f"[{source}] Row {data['row_num']} (Enroll: {enroll_no}): Invalid semester format '{sem_val}'")
                        error_count += 1
                        continue
                    
                    semester_obj, _ = Semester.objects.get_or_create(
                        number=int(sem_num),
                        defaults={'name': f'Semester {sem_num}'}
                    )
                    
                    username = enroll_no
                    user_defaults = {
                        'email': f'{username}@student.com',
                        'first_name': name.split()[0] if ' ' in name else name,
                        'last_name': ' '.join(name.split()[1:]) if ' ' in name else '',
                        'role': 'student',
                        'enrollment_no': enroll_no,
                        'is_first_login': True
                    }
                    
                    user, user_created = User.objects.get_or_create(
                        username=username,
                        defaults=user_defaults
                    )
                    
                    if user_created:
                        user.set_password(username)
                        user.save()
                        created_count += 1
                    elif update_existing:
                        user.first_name = user_defaults['first_name']
                        user.last_name = user_defaults['last_name']
                        user.enrollment_no = enroll_no
                        user.save()
                        updated_count += 1
                    else:
                        skipped_existing += 1
                    
                    StudentSemester.objects.update_or_create(
                        student=user,
                        session=session_obj,
                        defaults={
                            'branch': branch_obj,
                            'semester': semester_obj,
                            'class_name': semester_obj.year_name,
                            'is_active': True
                        }
                    )
                    processed_configs.add((session_obj, branch_obj, semester_obj))
                    
                except Exception as e:
                    db_errors.append(f"[{source}] Row {data['row_num']} (Enroll: {enroll_no}): Database error - {str(e)}")
                    error_count += 1
        
        for session_obj, branch_obj, semester_obj in processed_configs:
            has_offerings = SessionOffering.objects.filter(
                session=session_obj,
                base_offering__branch=branch_obj,
                base_offering__semester=semester_obj,
                is_active=True
            ).exists()
            
            if not has_offerings:
                warnings.append(f"Warning: No subject offerings found for {branch_obj.code} Semester {semester_obj.number} in '{session_obj.name}'.")

        return {
            'created': created_count,
            'updated': updated_count,
            'skipped_existing': skipped_existing,
            'errors': db_errors,
            'warnings': warnings,
            'total_processed': created_count + updated_count + skipped_existing
        }

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_students(request):
    user = request.user
    if user.role not in ['hod', 'admin']:
        raise PermissionDenied("Only HOD and Admin can upload students")
    
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    preview = request.data.get('preview', 'false').lower() == 'true'
    sheet_name = request.data.get('sheet_name', None)
    update_existing = request.data.get('update_existing', 'true').lower() == 'true'
    session_id = request.data.get('session_id', None)
    
    if not file.name.endswith(('.csv', '.xlsx', '.xls')):
        return Response({'error': 'Only CSV and Excel files are allowed'}, status=status.HTTP_400_BAD_REQUEST)
    
    result = StudentImportService.process(
        file, 
        user, 
        preview=preview, 
        sheet_name=sheet_name,
        update_existing=update_existing,
        session_id=session_id
    )
    
    if not result['success']:
        return Response(result, status=status.HTTP_400_BAD_REQUEST)
    
    return Response(result, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_student_upload_template(request):
    """
    Get template format for student upload as an Excel file using openpyxl
    """
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    
    headers = ['Name', 'Enroll Number', 'Department', 'Semester', 'Session']
    ws.append(headers)
    ws.append(['John Doe', '2024001', 'IT', '3', 'ODD 2024'])
    
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_upload_template.xlsx"'
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_session_students(request, session_id):
    user = request.user
    if user.role not in ['hod', 'admin']:
        raise PermissionDenied("Only HOD and Admin can view session students")
    
    session = get_object_or_404(FeedbackSession, pk=session_id)
    student_semesters = StudentSemester.objects.filter(
        session=session
    ).select_related('student', 'branch', 'semester').order_by('branch__name', 'semester__number', 'student__username')
    
    students_data = []
    for student_sem in student_semesters:
        students_data.append({
            'student_id': student_sem.student.id,
            'username': student_sem.student.username,
            'name': student_sem.student.get_full_name(),
            'enrollment_no': student_sem.student.enrollment_no,
            'branch': {
                'id': student_sem.branch.id,
                'name': student_sem.branch.name,
                'code': student_sem.branch.code
            },
            'semester': {
                'id': student_sem.semester.id,
                'number': student_sem.semester.number,
                'name': student_sem.semester.name
            },
            'class_name': student_sem.class_name,
            'roll_number': student_sem.roll_number,
            'is_active': student_sem.is_active,
            'created_at': student_sem.created_at
        })
    
    return Response({
        'session': FeedbackSessionSerializer(session).data,
        'students': students_data,
        'total_count': len(students_data)
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def assign_student_to_session(request):
    user = request.user
    if user.role not in ['hod', 'admin']:
        raise PermissionDenied("Only HOD and Admin can assign students")
    
    student_id = request.data.get('student_id')
    session_id = request.data.get('session_id')
    branch_id = request.data.get('branch_id')
    semester_id = request.data.get('semester_id')
    class_name = request.data.get('class_name', 'A')
    roll_number = request.data.get('roll_number', '')
    
    if not all([student_id, session_id, branch_id, semester_id]):
        return Response({'error': 'student_id, session_id, branch_id, and semester_id are required'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        student = get_object_or_404(User, pk=student_id, role='student')
        session = get_object_or_404(FeedbackSession, pk=session_id)
        branch = get_object_or_404(Branch, pk=branch_id)
        semester = get_object_or_404(Semester, pk=semester_id)
        
        if not class_name or class_name == 'A':
            class_name = semester.year_name

        student_semester, created = StudentSemester.objects.update_or_create(
            student=student,
            session=session,
            defaults={
                'branch': branch,
                'semester': semester,
                'class_name': class_name,
                'roll_number': roll_number,
                'is_active': True
            }
        )
        
        return Response({
            'message': 'Student assigned to session successfully',
            'student_semester': StudentSemesterSerializer(student_semester).data
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def remove_student_from_session(request, session_id, student_id):
    user = request.user
    if user.role not in ['hod', 'admin']:
        raise PermissionDenied("Only HOD and Admin can remove students")
    
    try:
        session = get_object_or_404(FeedbackSession, pk=session_id)
        student = get_object_or_404(User, pk=student_id, role='student')
        StudentSemester.objects.filter(student=student, session=session).delete()
        return Response({'message': 'Student removed from session successfully'})
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
